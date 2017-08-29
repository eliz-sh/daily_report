
# coding: utf-8

# In[32]:


import postgresql
import numpy as np
import postgresql.driver as pg_driver
import csv
import requests
import pandas as pd
import xlrd
import xlwt
import openpyxl
import psycopg2
from datetime import date, datetime, timedelta
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.cell import Cell


# In[33]:


#DB streetbee
db1 = pg_driver.connect(user='streetbee_analytics', password='ecmCZZNvLwAHk7onTVXC',host='pg-001-analysis.cnysugxwnx0x.eu-central-1.rds.amazonaws.com', database='streetbee', port=5432)

#Local DB
db2 = pg_driver.connect(user='elizaveta_sb', host='localhost', database='streetbee_analytics', port=5432)


# # Daily Report
#
# ## 1 Sheet

# In[34]:


sample_for_1sheeet=[]
sample_for_1sheeet = db1.query("WITH base_count_projects AS (SELECT task_projects.id, task_projects.wbs_code, COUNT(published_tasks.id) * max_response_count AS Base FROM public.published_tasks INNER JOIN public.task_projects ON task_projects.id=published_tasks.task_project_id WHERE published_tasks.state ='vacant' OR published_tasks.state ='completed' OR published_tasks.state ='reserved' GROUP BY task_projects.id) SELECT task_projects.id, task_projects.wbs_code, task_projects.deadline_at, Base FROM task_projects INNER JOIN base_count_projects ON base_count_projects.id = task_projects.id WHERE (task_projects.published_at <= now() AND task_projects.deadline_at > now()) AND (task_projects.id != 573)")

input_sheet1 = np.asarray(sample_for_1sheeet)

local_sample_1sheeet=[]
local_sample_1sheeet = db2.query("SELECT * FROM report_sheet1")

count_last_days = np.asarray(local_sample_1sheeet)

yesterday = date.today() - timedelta(1)
before_yesterday = date.today() - timedelta(2)
input_sheet1


# In[35]:


for i in range(11):
    def func(a):
        return np.append(a, 0)
    input_sheet1 = np.apply_along_axis(func, 1, input_sheet1)

for x in range(len(count_last_days)):
    if before_yesterday==count_last_days[x][2]:
        for i in range(len(input_sheet1)):
            if count_last_days[x][0]==input_sheet1[i][0]:
                input_sheet1[i][4]=count_last_days[x][1]

    if yesterday==count_last_days[x][2]:
        for i in range(len(input_sheet1)):
            if count_last_days[x][0]==input_sheet1[i][0]:
                input_sheet1[i][5]=count_last_days[x][1]

    if date.today()==count_last_days[x][2]:
        for i in range(len(input_sheet1)):
            if count_last_days[x][0]==input_sheet1[i][0]:
                input_sheet1[i][6]=count_last_days[x][1]


# In[36]:


#группировка по wbs
for i in range(len(input_sheet1)):
    for j in range(len(input_sheet1)):
        if input_sheet1[i][1] == input_sheet1[j][1] and i!=j and input_sheet1[i][1]!=0 and input_sheet1[j][1]!=0:
            input_sheet1[i][0] = str(input_sheet1[i][0])+ "," + str(input_sheet1[j][0])
            input_sheet1[i][3] += input_sheet1[j][3]
            input_sheet1[i][4] += input_sheet1[j][4]
            input_sheet1[i][5] += input_sheet1[j][5]
            input_sheet1[i][6] += input_sheet1[j][6]
            input_sheet1[j]=0

sheet1=[]
for row in input_sheet1:
    if row[0]!=0:
        sheet1 = np.append(sheet1, row)

sheet1=np.reshape(sheet1, ((len(sheet1)//15),15))


# In[37]:


for i in range(len(sheet1)):
    sheet1[i][2]=sheet1[i][2].date()   #convert datetime to date
    for x in range(4, 7):
        sheet1[i][x+3]=sheet1[i][x]/sheet1[i][3] #Count percent the last three days
    sheet1[i][10]=(sheet1[i][9]-sheet1[i][7])/3 #Speed (%perday)==AvarageSpeed
    sheet1[i][11]=sheet1[i][6]-sheet1[i][5] #1delta
    if sheet1[i][10]!=0:
        sheet1[i][12] = date.today() + timedelta((1-sheet1[i][9])/sheet1[i][10]) #4Cast Finish
    if sheet1[i][12]!=0:
        sheet1[i][13]= (sheet1[i][12] - sheet1[i][2]).days #Delay
    sheet1[i][14]= ((sheet1[i][2] - date.today()).days)*sheet1[i][10]+sheet1[i][9]   #Done
    if sheet1[i][14] > 100:
        sheet1[i][14]=100

bef_ye=str(before_yesterday)+'(%)'
yest=str(yesterday)+'(%)'
today=str(date.today())+'(%)'

# for i in range(len(sheet1)):
#     sheet1[i][7]="{0:.0f}%".format(sheet1[i][7])
#     sheet1[i][8]="{0:.0f}%".format(sheet1[i][8])
#     sheet1[i][9]="{0:.0f}%".format(sheet1[i][9])

sheet1 = pd.DataFrame(sheet1, columns=['id', 'wbs_code', 'Deadline', 'Base', before_yesterday, yesterday, date.today(), bef_ye, yest, today , 'Average speed', 'Delta 1 Day', 'Expected Completion', 'Delay', 'Done (%)'], )
sheet1

# print "{0:.0f}%".format(1./3 * 100)


# ## Sheet 2

# In[38]:


CSV_URL = 'https://docs.google.com/spreadsheets/d/103IXtgwcdMbUMCtDLT7Hs-GnMX0sJgxBSPv8rxPuK_U/export?format=csv&gid=0'

with requests.Session() as s:
    download = s.get(CSV_URL)

    decoded_content = download.content.decode('utf-8')

    cr = csv.reader(decoded_content.splitlines(), delimiter=',')
    my_list = list(cr)

input_coordinator = pd.DataFrame(my_list, columns=['coordinator','region'])


# In[39]:


# completed
sample_for_completed=[]
sample_for_completed = db1.query("""
WITH active_projects AS (SELECT  DISTINCT wbs_code FROM task_projects
WHERE published_at <= now() AND deadline_at > now() AND (task_projects.id != 573))
SELECT task_points.region, task_points.city, task_projects.wbs_code, Count(*) AS Completed
FROM reserved_tasks INNER JOIN published_tasks ON published_tasks.id = reserved_tasks.published_task_id
INNER JOIN task_points ON task_points.id = published_tasks.task_point_id
INNER JOIN task_projects ON task_projects.id = published_tasks.task_project_id
WHERE task_projects.wbs_code IN (SELECT wbs_code FROM active_projects)
AND (reserved_tasks.state = 'response_received' OR reserved_tasks.state = 'response_accepted')
AND published_tasks.type = 'PublishedTasks::PointTask'
GROUP BY task_points.region, task_projects.wbs_code, task_points.city ORDER BY task_points.region, city;""")

input_completed = pd.DataFrame(sample_for_completed, columns=['region','city', 'wbs code','Completed'])
input_completed=pd.merge(input_coordinator, input_completed, on="region")

part1=pd.pivot_table(input_completed,index=['coordinator','region', "city"], columns=["wbs code"] , fill_value=0)
part2=pd.pivot_table(input_completed,index=['coordinator','region',"city"], values=["Completed"], aggfunc={np.sum} , fill_value=0)
part2=part2.rename(columns={'sum': 'Итого'})
completed=pd.concat([part1, part2],axis=1 )

# Completed % of Total (Base)
sample_for_percent=[]
sample_for_percent= db1.query("""
WITH active_projects AS
  (SELECT DISTINCT wbs_code FROM task_projects WHERE published_at <= now() AND deadline_at > now()  AND (task_projects.id != 573))
SELECT task_points.region, task_points.city, task_projects.wbs_code, Count(*) AS Completed
FROM published_tasks
  INNER JOIN task_points ON task_points.id = published_tasks.task_point_id
  INNER JOIN task_projects ON task_projects.id = published_tasks.task_project_id
WHERE task_projects.wbs_code IN (SELECT active_projects.wbs_code FROM active_projects)
      AND published_tasks.state IN ('vacant', 'completed', 'reserved')
      AND published_tasks.type = 'PublishedTasks::PointTask'
GROUP BY task_points.region, task_projects.wbs_code, task_points.city ORDER BY task_points.region, city;""")

input_total = pd.DataFrame(sample_for_percent, columns=['region','city', 'wbs code','Completed'])
input_total=pd.merge(input_coordinator, input_total, on="region")

part1=pd.pivot_table(input_total,index=['coordinator','region', "city"],columns=["wbs code"] , fill_value=0)
part2=pd.pivot_table(input_total,index=['coordinator','region', "city"], values=["Completed"], aggfunc={np.sum} ,fill_value=0)
part2=part2.rename(columns={'sum': 'Итого'})

Base=part2.rename(columns={'Completed': 'Base'})

total=pd.concat([part1, part2],axis=1)

completed_percent=completed.divide(total, axis=0, level=None, fill_value=0)
# completed_percent=completed_percent*100
completed_percent=completed_percent.fillna(0)
completed_percent=completed_percent.rename(columns={'Completed': 'Completed % of Base'})
# completed_percent=completed_percent.astype(float)

a = completed_percent.merge(completed, left_index=True, right_index=True, how='outer')
a


# In[40]:


# Delta 1 day
yesterday=db2.query("SELECT * FROM report_sheet2 WHERE date=(current_date - interval '1 day');")

delta1=[]
for row in sample_for_completed:
    for rownum in yesterday:
        if row[0:3] == rownum[0:3]:
            delta1=np.append(delta1, [row[0],row[1], row[2], row[3]-rownum[3]])
l=len(delta1)//4
delta1=np.reshape(delta1, (l,4))
for row in delta1:
    row[3]=int(row[3])
delta1=pd.DataFrame(delta1,columns=['region','city', 'wbs code','Delta1'])
delta1=pd.merge(input_coordinator, delta1, on="region")

part1=pd.pivot_table(delta1,index=['coordinator','region',"city"],columns=["wbs code"] , aggfunc={np.sum}, fill_value=0)
part2=pd.pivot_table(delta1,index=['coordinator','region',"city"], values=["Delta1"], aggfunc={np.sum} ,fill_value=0)
part1=part1['Delta1']
part1=part1.rename(columns={'sum': 'Delta1'})
part2=part2.rename(columns={'sum': 'Итого'})
Delta1=pd.concat([part1, part2],axis=1)


b = a.merge(Delta1, left_index=True, right_index=True, how='outer')

# Delta 3 day
three_days_ago=db2.query("SELECT * FROM report_sheet2 WHERE date=(current_date - interval '3 day');")

delta3=[]
for row in sample_for_completed:
    for rownum in three_days_ago:
        if row[0:3] == rownum[0:3]:
            delta3=np.append(delta3, [row[0],row[1], row[2], row[3]-rownum[3]])
l=len(delta3)//4
delta3=np.reshape(delta3, (l,4))
for row in delta3:
    row[3]=int(row[3])
delta3=pd.DataFrame(delta3,columns=['region','city', 'wbs code','Delta3'])
delta3=pd.merge(input_coordinator, delta3, on="region")
print(delta3)
part1=pd.pivot_table(delta3,index=['coordinator','region',"city"],columns=["wbs code"] , aggfunc={np.sum}, fill_value=0)
part2=pd.pivot_table(delta3,index=['coordinator','region',"city"], values=["Delta3"], aggfunc={np.sum} ,fill_value=0)
part1=part1['Delta3']
part1=part1.rename(columns={'sum': 'Delta3'})
part2=part2.rename(columns={'sum': 'Итого'})
Delta3=pd.concat([part1, part2],axis=1)

c = b.merge(Delta3, left_index=True, right_index=True, how='outer')

#  Delta 7 days
seven_days_ago=db2.query("SELECT * FROM report_sheet2 WHERE date=(current_date - interval '7 day');")

delta7=[]
for row in sample_for_completed:
    for rownum in seven_days_ago:
        if row[0:3] == rownum[0:3]:
            delta7=np.append(delta7, [row[0],row[1], row[2], row[3]-rownum[3]])

l=len(delta7)//4
delta7=np.reshape(delta7, (l,4))
for row in delta7:
    row[3]=int(row[3])
delta7=pd.DataFrame(delta7,columns=['region','city', 'wbs code','Delta7'])
delta7=pd.merge(input_coordinator, delta7, on="region")

part1=pd.pivot_table(delta7,index=['coordinator','region',"city"],columns=["wbs code"] , aggfunc={np.sum}, fill_value=0)
part2=pd.pivot_table(delta7,index=['coordinator','region',"city"], values=["Delta7"], aggfunc={np.sum} ,fill_value=0)
part1=part1['Delta7']
part1=part1.rename(columns={'sum': 'Delta7'})
part2=part2.rename(columns={'sum': 'Итого'})
Delta7=pd.concat([part1, part2],axis=1)

d = c.merge(Delta7, left_index=True, right_index=True, how='outer')

# d = b.merge(Delta7, left_index=True, right_index=True, how='outer')


# In[41]:


# vacant
sample_vacant=[]
sample_vacant = db1.query("""
WITH active_projects AS (SELECT DISTINCT wbs_code  FROM task_projects WHERE published_at <= now() AND deadline_at > now()
                                                                      AND (task_projects.id != 573))
SELECT task_points.region, task_points.city, task_projects.wbs_code, Count(*) AS Vacant FROM published_tasks
  INNER JOIN task_points ON task_points.id = published_tasks.task_point_id
  INNER JOIN task_projects ON task_projects.id = published_tasks.task_project_id
WHERE task_projects.wbs_code IN (SELECT active_projects.wbs_code FROM active_projects) AND published_tasks.state = 'vacant' AND published_tasks.type = 'PublishedTasks::PointTask'
GROUP BY task_points.region, task_projects.wbs_code, task_points.city ORDER BY task_points.region, city;
""")
input_vacant = pd.DataFrame(sample_vacant, columns=['region','city', 'wbs code','Vacant'])
input_vacant=pd.merge(input_coordinator, input_vacant, on="region")

part1=pd.pivot_table(input_vacant,index=['coordinator','region',"city"], columns=["wbs code"] , fill_value=0)
part2=pd.pivot_table(input_vacant,index=['coordinator','region',"city"], values=["Vacant"], aggfunc={np.sum} , fill_value=0)
part2=part2.rename(columns={'sum': 'Итого'})
vacant=pd.concat([part1, part2],axis=1 )
# vacant.head().sort(by=['Vacant', ["Итого"]],ascending=False)

e = d.merge(vacant, left_index=True, right_index=True, how='outer')

# active
sample_active=[]
sample_active = db1.query("""
WITH active_projects AS
  (SELECT DISTINCT wbs_code FROM task_projects WHERE published_at <= now() AND deadline_at > now()  AND (task_projects.id != 573))
SELECT task_points.region, task_points.city, task_projects.wbs_code, Count(*) AS Vacant FROM reserved_tasks
  INNER JOIN published_tasks ON published_tasks.id = reserved_tasks.published_task_id
  INNER JOIN task_points ON task_points.id = published_tasks.task_point_id
  INNER JOIN task_projects ON task_projects.id = published_tasks.task_project_id
WHERE task_projects.wbs_code IN (SELECT active_projects.wbs_code FROM active_projects) AND reserved_tasks.state = 'active' AND published_tasks.type = 'PublishedTasks::PointTask'
GROUP BY task_points.region, task_projects.wbs_code, task_points.city ORDER BY task_points.region, city;
                       """  )
input_active = pd.DataFrame(sample_active, columns=['region','city', 'wbs code','Active'])
input_active=pd.merge(input_coordinator, input_active, on="region")

part1=pd.pivot_table(input_active,index=['coordinator','region',"city"], columns=["wbs code"] , fill_value=0)
part2=pd.pivot_table(input_active,index=['coordinator','region',"city"], values=["Active"], aggfunc={np.sum} , fill_value=0)
part2=part2.rename(columns={'sum': 'Итого'})
active=pd.concat([part1, part2],axis=1 )

f = e.merge(active, left_index=True, right_index=True, how='outer')
sheet2=f.merge(Base, left_index=True, right_index=True, how='outer')
sheet2=sheet2.fillna(0)


sheet2.columns


# ## Sheet 3

# In[42]:


# CSV_URL = 'https://docs.google.com/spreadsheets/d/1YcWRNcZu2J0ZgU8jQpfv7tzI6QlR4hgc3acb4htOs3A/export?format=csv&gid=0'
CSV_URL = 'https://docs.google.com/spreadsheets/d/19YfEi8TUWUzCqRmlpiOcR1toI3Tr0SVRbQwqHg1wv9s/export?format=csv&gid=721464937'

with requests.Session() as s:
    download = s.get(CSV_URL)

    decoded_content = download.content.decode('utf-8')

    cr = csv.reader(decoded_content.splitlines(), delimiter=',')
    my_list = list(cr)
#     for row in my_list:
#         row[1]=int(row[1])

my_list=pd.DataFrame(my_list)

my_list=my_list.drop([0,1], axis=0)
input_analytics=my_list.iloc[:,[0,10]]
input_analytics=input_analytics.rename(columns={0:'wbs code' , 10:'analyst'})
# input_analytics


# In[43]:


sample_received=[]
sample_received = db1.query("WITH active_projects AS (SELECT id, wbs_code FROM task_projects WHERE published_at <= now() AND deadline_at > now()  AND (task_projects.id != 573)) SELECT *, case WHEN interval_of_days < 1 then 1 WHEN interval_of_days < 3 then 2 ELSE 3 end as waiting_class FROM (SELECT task_projects.wbs_code, (extract(EPOCH FROM now() - response_sets.created_at) / 3600 / 24) :: INT AS interval_of_days FROM response_sets INNER JOIN reserved_tasks ON reserved_tasks.id = response_sets.reserved_task_id INNER JOIN published_tasks ON published_tasks.id = reserved_tasks.published_task_id INNER JOIN task_projects ON task_projects.id = published_tasks.task_project_id WHERE task_project_id IN (SELECT id FROM active_projects) AND reserved_tasks.state = 'response_received') t1 ORDER BY waiting_class")

sheet3 = pd.DataFrame(sample_received, columns=['wbs code','days' ,'interval'])
sheet3 = pd.merge(input_analytics, sheet3, on="wbs code")
# sheet3
part1=pd.pivot_table(sheet3,index=['analyst',"wbs code"],columns=["interval"],aggfunc={np.count_nonzero} , fill_value=0)
part1=part1['days']['count_nonzero']

part1=part1.rename(columns={1: '<24', 2: '24-48', 3: '48+'})
part2=pd.pivot_table(sheet3,index=['analyst',"wbs code"], values=['days'], aggfunc={np.count_nonzero} , fill_value=0)
part2=part2['days']
part2=part2.rename(columns={'count_nonzero': 'Итого'})
sheet3=pd.concat([part1, part2],axis=1 )
sheet3


# In[44]:


len_completed_percent=len(completed_percent.columns)
len_completed=len(completed.columns)
len_delta1=len(Delta1.columns)
len_delta3=len(Delta3.columns)
len_delta7=len(Delta7.columns)
len_vacant=len(vacant.columns)
len_active=len(active.columns)


# In[56]:


str(date.today())


# # Вывод в Excel

# In[58]:


sh1 = pd.DataFrame(sheet1)
sh2 = pd.DataFrame(sheet2)
sh3 = pd.DataFrame(sheet3)
d=str(date.today())
s = "";
seq = ("Daily_Report_",d,".xlsx"); # This is sequence of strings.
c = s.join( seq )
# c="Daily_Report "+str(today())+".xlsx"
writer = pd.ExcelWriter(c, engine='xlsxwriter')


# Convert the dataframe to an XlsxWriter Excel object.
sh1.to_excel(writer, sheet_name='Общий', index=False)
sh2.to_excel(writer, sheet_name='Территории',merge_cells=False,header=False,startrow=2)
sh3.to_excel(writer, sheet_name='Received')

# # Get the xlsxwriter workbook and worksheet objects for 1 Sheet
workbook  = writer.book
worksheet1 = writer.sheets['Общий']

format_percent = workbook.add_format({'num_format': '0%'})

worksheet1.set_row(0, 35)
worksheet1.set_column('B:B',25)
worksheet1.set_column('C:C',12)
worksheet1.set_column('D:D',8)
worksheet1.set_column('E:G',12)
worksheet1.set_column('H:K',12, format_percent)
worksheet1.set_column('L:N',10)
worksheet1.set_column('O:O',10, format_percent)

df_headers= pd.DataFrame([], [], sh2.columns)
df_headers.to_excel(writer, sheet_name='Территории', merge_cells=True, startcol=2)

worksheet2 = writer.sheets['Территории']
worksheet2.set_column('A:A',20)
worksheet2.set_column('B:B',25)
worksheet2.set_column('C:C',20)
worksheet2.set_row(1, 120)

for i in range (4, 3+len_completed_percent):
    worksheet2.set_column(i,i, 5, format_percent)

ll= 3+len_completed_percent+len_completed+len_delta1+len_delta3+len_delta7+len_vacant+len_active

for i in range(3+len_completed_percent,ll):
    worksheet2.set_column(i,i,4)

worksheet3 = writer.sheets['Received']
worksheet3.set_column('A:A',15)
worksheet3.set_column('B:B',25)

writer.save()


# In[59]:


wb = load_workbook(c)

ws1 = wb.get_sheet_by_name('Общий')
ws = wb.get_sheet_by_name('Территории')
ws3 = wb.get_sheet_by_name('Received')

for col in ws1.iter_cols(max_col=15):
    for cell in col:
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        cell.alignment = Alignment(wrapText=True)



for col in ws1.iter_cols(min_col=1, max_col=4, max_row=1):
    for cell in col:
        cell.fill = PatternFill(start_color= 'FDE9D9', end_color= 'ff0000', fill_type = "solid")

for col in ws1.iter_cols(min_col=5, max_col=7,max_row=1):
    for cell in col:
        cell.fill = PatternFill(start_color= 'FCD5B4', end_color= 'ff0000', fill_type = "solid")

for col in ws1.iter_cols(min_col=8, max_col=10, max_row=1):
    for cell in col:
        cell.fill = PatternFill(start_color= 'C4D79B', end_color= 'ff0000', fill_type = "solid")

for col in ws1.iter_cols(min_col=11, max_col=15,max_row=1):
    for cell in col:
        cell.fill = PatternFill(start_color= '9BBB59', end_color= 'ff0000', fill_type = "solid")

for col in ws3.iter_cols(max_col=6):
    for cell in col:
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(bold=False)
        cell.border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for col in ws3.iter_cols(max_col=6, min_row=1, max_row=1):
    for cell in col:
        cell.fill = PatternFill(start_color= '66c0b7', end_color= 'ff0000', fill_type = "solid")

for col in ws.columns:
    cell_A = col[1]
    cell_A.alignment = Alignment(text_rotation=90)

currentCell = ws['A2']
currentCell.alignment = Alignment(text_rotation=0)
currentCell = ws['B2'] #or currentCell = ws['A1']
currentCell.alignment = Alignment(text_rotation=0)
currentCell = ws['C2'] #or currentCell = ws['A1']
currentCell.alignment = Alignment(text_rotation=0)
ws['A2']= 'coordinator'
ws['B2']= 'region'
ws['C2']= 'city'

for col in ws.iter_cols(min_col=3, max_col=3, max_row=2):
    for cell in col:
        cell.border = Border(left=Side(style=None),
                     top=Side(style=None))


#coloring of head
m=4+len_completed_percent
for col in ws.iter_cols(min_col=1, max_col=m-1, max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= 'FDE9D9', end_color= 'ff0000', fill_type = "solid")

m1=m+len_completed
for col in ws.iter_cols(min_col=m, max_col=m1-1,max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= 'FCD5B4', end_color= 'ff0000', fill_type = "solid")

m2=m1+len_delta1
for col in ws.iter_cols(min_col=m1, max_col=m2-1, max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= 'C4D79B', end_color= 'ff0000', fill_type = "solid")

m3=m2+len_delta3
for col in ws.iter_cols(min_col=m2, max_col=m3-1,max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= '9BBB59', end_color= 'ff0000', fill_type = "solid")

m4=m3+len_delta7
for col in ws.iter_cols(min_col=m3, max_col=m4-1,max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= '76933C', end_color= 'ff0000', fill_type = "solid")

m5=m4+len_vacant
for col in ws.iter_cols(min_col=m4, max_col=m5-1, max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= 'DCE6F1', end_color= 'ff0000', fill_type = "solid")

m6=m5+len_active
for col in ws.iter_cols(min_col=m5, max_col=m6-1,max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= 'B8CCE4', end_color= 'ff0000', fill_type = "solid")

for col in ws.iter_cols(min_col=m6, max_col=m6, max_row=2):
    for cell in col:
        cell.fill = PatternFill(start_color= '31869B', end_color= 'ff0000', fill_type = "solid")

for col in ws.iter_cols(max_col=3):
    for cell in col:
        cell.alignment = Alignment(horizontal='left')
        cell.font = Font(bold=False)


for col in ws.iter_cols(min_col=4, max_col=m6, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='hair', color="ffffff"),
                     right=Side(style='hair', color='ffffff'),
                     top=Side(style='hair', color='ffffff'),
                     bottom=Side(style='hair',color='ffffff'))

#Border for blocks
for col in ws.iter_cols(min_col=4+len_completed_percent, max_col=4+len_completed_percent, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #right of Completed
for col in ws.iter_cols(min_col=m1, max_col=m1, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #right of Completed %
for col in ws.iter_cols(min_col=m2, max_col=m2, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #right of D1
for col in ws.iter_cols(min_col=m3, max_col=m3, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #right of D3
for col in ws.iter_cols(min_col=m4, max_col=m4, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #left of vacant
for col in ws.iter_cols(min_col=m5, max_col=m5, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')) #left of active
for col in ws.iter_cols(min_col=m6, max_col=m6, min_row=3):
    for cell in col:
        cell.border = Border(left=Side(style='thin'),top=Side(style='hair', color='ffffff')
                            ,right=Side(style='thin')) #base


for col in ws.iter_cols(min_col=4,min_row=3):
    for cell in col:
        if cell.value==0:
            cell.value='-'



wb.save(c)
